from fastapi import APIRouter

from database import (
    products_collection,
    orders_collection,
    warehouse_collection
)
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font
import os

router = APIRouter()


@router.get("/dashboard")
async def dashboard():
    total_products = await products_collection.count_documents({})
    total_orders = await orders_collection.count_documents({})

    orders = await orders_collection.find().to_list(100)

    revenue = sum(o.get("total", 0) for o in orders)


    global_low_stock = []

    products = await products_collection.find().to_list(100)

    for p in products:

        for v in p.get("variants", []):

            stock = v.get("stock", 0)

            if stock < 5:

                # LOW STOCK STATUS
                if stock <= 2:
                    status = "Critical"
                elif stock <= 5:
                    status = "Low"
                else:
                    status = "Warning"

                global_low_stock.append({
                    "product": p.get("name"),
                    "sku": v.get("sku"),
                    "stock": stock,
                    "status": status,
                    "reorderLevel": 10,
                    "suggestedReorder": 20 - stock
                })

    low_stock = 0
    low_stock_items = []

    warehouse_summary = {}

    warehouse_performance = []

    warehouses = await warehouse_collection.find().to_list(100)

    for wh in warehouses:

        warehouse_name = wh.get("name", "Unknown Warehouse")

        total_stock = 0
        inventory_value = 0
        low_items = 0

        for item in wh.get("inventory", []):

            sku = item.get("sku")
            qty = item.get("quantity", 0)

            total_stock += qty

            product = await products_collection.find_one({
                "variants.sku": sku
            })
            product_name = "Unknown Product"
            price = 0

            if product:
                product_name = product.get("name")

                for v in product.get("variants", []):
                    if v.get("sku") == sku:
                        price = v.get("price", 0)
                        break

            inventory_value += qty * price

            if qty < 5:

                low_stock += 1
                low_items += 1

                # STATUS
                if qty <= 2:
                    status = "Critical"
                elif qty <= 5:
                    status = "Low"
                else:
                    status = "Warning"

                low_stock_items.append({
                    "product": product_name,
                    "sku": sku,
                    "stock": qty,
                    "warehouse": warehouse_name,
                    "status": status,
                    "reorderLevel": 10,
                    "suggestedReorder": 20 - qty
                })
        warehouse_summary[warehouse_name] = {
            "stock": total_stock,
            "value": inventory_value
        }

        warehouse_performance.append({
            "warehouse": warehouse_name,
            "inventory": total_stock,
            "lowStockItems": low_items,
            "inventoryValue": inventory_value
        })

        recent_orders = [
        {
            "id": str(o["_id"]),

            "customer": (
                o.get("viewer_name")
                or o.get("viewer_email")
                or o.get("user_name")
                or o.get("user_email")
                or o.get("email")
                or "Unknown"
            ),

            "total": o.get("total", 0)
        }

        for o in orders[:5]
    ]
    warehouse_data = [
        {
            "warehouse": name,
            "stock": data["stock"],
            "value": data["value"]
        }

        for name, data in warehouse_summary.items()
    ]
        # ================= TOP SELLING PRODUCTS =================

    sales_map = {}

    for order in orders:

        for item in order.get("items", []):

            sku = item.get("sku")
            qty = item.get("quantity", 1)

            name = (
                item.get("name")
                or item.get("product_name")
                or "Unknown Product"
            )

            if sku not in sales_map:

                sales_map[sku] = {
                    "product": name,
                    "sku": sku,
                    "sold": 0,
                    "revenue": 0
                }

            sales_map[sku]["sold"] += qty

            sales_map[sku]["revenue"] += (
                item.get("price", 0) * qty
            )

    top_selling_products = sorted(
        sales_map.values(),
        key=lambda x: x["sold"],
        reverse=True
    )[:5]


    supplier_map = {}

    for p in products:

        supplier = (
            p.get("supplier_name")
            or "Unknown Supplier"
        )

        total_products_supplier = len(
            p.get("variants", [])
        )

        total_stock_supplier = sum(
            v.get("stock", 0)
            for v in p.get("variants", [])
        )

        inventory_value = sum(
            (
                v.get("stock", 0)
                * v.get("price", 0)
            )
            for v in p.get("variants", [])
        )

        if supplier not in supplier_map:

            supplier_map[supplier] = {
                "supplier": supplier,
                "products": 0,
                "stock": 0,
                "value": 0
            }

        supplier_map[supplier]["products"] += (
            total_products_supplier
        )

        supplier_map[supplier]["stock"] += (
            total_stock_supplier
        )

        supplier_map[supplier]["value"] += (
            inventory_value
        )

    supplier_report = list(
        supplier_map.values()
    )


    dead_stock = []

    sold_skus = set()

    for order in orders:

        for item in order.get("items", []):

            sold_skus.add(
                item.get("sku")
            )

    for p in products:

        for v in p.get("variants", []):

            sku = v.get("sku")

            if sku not in sold_skus:

                dead_stock.append({

                    "product": p.get("name"),

                    "sku": sku,

                    "stock": v.get("stock", 0),

                    "price": v.get("price", 0),

                    "inventoryValue": (
                        v.get("stock", 0)
                        * v.get("price", 0)
                    )
                })

    return {

        "products": total_products,
        "orders": total_orders,
        "revenue": revenue,

        "lowStock": low_stock,

        "lowStockReport": low_stock_items,
        "warehouseSummary": warehouse_data,
        "globalLowStock": global_low_stock,
        "warehousePerformance": warehouse_performance,
        "topSellingProducts": top_selling_products,

        "supplierReport": supplier_report,

        "deadStockReport": dead_stock,
    }




@router.get("/dashboard/export/excel")
async def export_dashboard_excel():

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Low Stock Report"

    headers = [
        "Product",
        "SKU",
        "Stock",
        "Status",
        "Reorder Level",
        "Suggested Reorder"
    ]

    ws1.append(headers)

    for cell in ws1[1]:
        cell.font = Font(bold=True)

    products = await products_collection.find().to_list(100)

    for p in products:

        for v in p.get("variants", []):

            stock = v.get("stock", 0)

            if stock < 5:

                status = "Low"

                if stock <= 2:
                    status = "Critical"

                ws1.append([
                    p.get("name"),
                    v.get("sku"),
                    stock,
                    status,
                    10,
                    20 - stock
                ])



    ws2 = wb.create_sheet(title="Top Selling")

    ws2.append([
        "Product",
        "SKU",
        "Units Sold"
    ])

    for cell in ws2[1]:
        cell.font = Font(bold=True)

    sales_map = {}

    orders = await orders_collection.find().to_list(100)

    for order in orders:

        for item in order.get("items", []):

            sku = item.get("sku")
            qty = item.get("quantity", 1)

            name = (
                item.get("name")
                or item.get("product_name")
                or "Unknown Product"
            )

            if sku not in sales_map:
                sales_map[sku] = {
                    "product": name,
                    "sold": 0
                }

            sales_map[sku]["sold"] += qty

    sorted_sales = sorted(
        sales_map.items(),
        key=lambda x: x[1]["sold"],
        reverse=True
    )

    for sku, data in sorted_sales:

        ws2.append([
            data["product"],
            sku,
            data["sold"]
        ])


    ws3 = wb.create_sheet(title="Warehouse Summary")

    ws3.append([
        "Warehouse",
        "Total Units",
        "Inventory Value"
    ])

    for cell in ws3[1]:
        cell.font = Font(bold=True)

    warehouses = await warehouse_collection.find().to_list(100)

    for wh in warehouses:

        total_stock = 0
        total_value = 0

        for item in wh.get("inventory", []):

            qty = item.get("quantity", 0)
            total_stock += qty

            product = await products_collection.find_one({
                "variants.sku": item.get("sku")
            })

            price = 0

            if product:
                for v in product.get("variants", []):
                    if v.get("sku") == item.get("sku"):
                        price = v.get("price", 0)
                        break

            total_value += qty * price

        ws3.append([
            wh.get("name"),
            total_stock,
            total_value
        ])


    ws4 = wb.create_sheet(title="Supplier Report")

    ws4.append([
        "Supplier",
        "Products",
        "Stock",
        "Inventory Value"
    ])

    for cell in ws4[1]:
        cell.font = Font(bold=True)

    supplier_map = {}

    for p in products:

        supplier = (
            p.get("supplier_name")
            or "Unknown Supplier"
        )

        if supplier not in supplier_map:
            supplier_map[supplier] = {
                "products": 0,
                "stock": 0,
                "value": 0
            }

        supplier_map[supplier]["products"] += 1

        for v in p.get("variants", []):

            stock = v.get("stock", 0)
            price = v.get("price", 0)

            supplier_map[supplier]["stock"] += stock
            supplier_map[supplier]["value"] += stock * price

    for supplier, data in supplier_map.items():

        ws4.append([
            supplier,
            data["products"],
            data["stock"],
            data["value"]
        ])


    ws5 = wb.create_sheet(title="Dead Stock")

    ws5.append([
        "Product",
        "SKU",
        "Stock",
        "Status"
    ])

    for cell in ws5[1]:
        cell.font = Font(bold=True)

    for p in products:

        for v in p.get("variants", []):

            stock = v.get("stock", 0)

            if stock > 0:

                sku = v.get("sku")

                sold = False

                for order in orders:

                    for item in order.get("items", []):

                        if item.get("sku") == sku:
                            sold = True
                            break

                    if sold:
                        break

                if not sold:

                    ws5.append([
                        p.get("name"),
                        sku,
                        stock,
                        "Dead Stock"
                    ])

    os.makedirs("exports", exist_ok=True)

    file_path = "exports/dashboard_report.xlsx"

    wb.save(file_path)

    return FileResponse(
        path=file_path,
        filename="dashboard_report.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )